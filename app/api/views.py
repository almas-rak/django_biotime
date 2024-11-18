import re

from django.http import HttpResponse
from rest_framework import status
from rest_framework.views import APIView
from rest_framework.response import Response
from rest_framework.permissions import IsAuthenticated
from datetime import datetime, timedelta

import openpyxl
from openpyxl.styles import Font, Alignment, Border, Side, PatternFill
from openpyxl.utils import get_column_letter
from io import BytesIO

from zk_services.views import ZkApiInteraction


class GetEmp(APIView):
    permission_classes = (IsAuthenticated,)

    def get(self, request):
        zk_api_interaction = ZkApiInteraction()
        zk_api_interaction.refresh_token()
        if isinstance(zk_api_interaction.token, dict) and 'status_code' in zk_api_interaction.token:
            content = zk_api_interaction.token
        else:
            emp = zk_api_interaction.get_emp(user=request.user)
            content = {'data': emp}
        return Response(content)


class GetMonthlyStatusReport(APIView):
    permission_classes = (IsAuthenticated,)

    def get(self, request):
        start_date = request.query_params.get("start_date")
        end_date = request.query_params.get("end_date")
        employees = request.query_params.get("employees")
        if not start_date:
            return Response({"error": f"Нет параметра: start_date"}, status=status.HTTP_400_BAD_REQUEST)
        if not end_date:
            return Response({"error": f"Нет параметра: end_date"}, status=status.HTTP_400_BAD_REQUEST)
        if not employees:
            return Response({"error": f"Нет параметра: employees"}, status=status.HTTP_400_BAD_REQUEST)
        zk_api_interaction = ZkApiInteraction()
        zk_api_interaction.refresh_token()
        if isinstance(zk_api_interaction.token, dict) and 'status_code' in zk_api_interaction.token:
            content = zk_api_interaction.token
        else:
            params = {
                "start_date": start_date,
                "end_date": end_date,
                "employees": employees,
                "page_size": 100
            }
            status_report = zk_api_interaction.get_monthly_status_report(params=params, user=request.user)
            content = status_report
        return Response(content)


class GetMonthlyPunchReport(APIView):
    permission_classes = (IsAuthenticated,)

    def get(self, request):
        start_date = request.query_params.get("start_date")
        end_date = request.query_params.get("end_date")
        employees = request.query_params.get("employees")
        if not start_date:
            return Response({"error": f"Нет параметра: start_date"}, status=status.HTTP_400_BAD_REQUEST)
        if not end_date:
            return Response({"error": f"Нет параметра: end_date"}, status=status.HTTP_400_BAD_REQUEST)
        if not employees:
            return Response({"error": f"Нет параметра: employees"}, status=status.HTTP_400_BAD_REQUEST)
        zk_api_interaction = ZkApiInteraction()
        zk_api_interaction.refresh_token()
        if isinstance(zk_api_interaction.token, dict) and 'status_code' in zk_api_interaction.token:
            content = zk_api_interaction.token
        else:
            params = {
                "start_date": start_date,
                "end_date": end_date,
                "employees": employees,
                "page_size": 100
            }
            punch_report = zk_api_interaction.get_monthly_punch_report(params=params, user=request.user)
            content = punch_report
        return Response(content)


class EmpSearch(APIView):
    permission_classes = (IsAuthenticated,)

    def get(self, request):
        params = {
            "employee_icontains": request.query_params.get("employee_icontains"),
            "page_size": 100
        }
        zk_api_interaction = ZkApiInteraction()
        zk_api_interaction.refresh_token()
        if isinstance(zk_api_interaction.token, dict) and 'status_code' in zk_api_interaction.token:
            content = zk_api_interaction.token
        else:
            content = zk_api_interaction.emp_search(params=params, user=request.user)

        return Response(content)


def timedelta_hours_minutes_str(tot_time):
    total_seconds = int(tot_time.total_seconds())
    hours, remainder = divmod(total_seconds, 3600)
    minutes, _ = divmod(remainder, 60)
    return f"{hours}:{minutes:02d}"


def pars_data(status_report, punch_report):
    data = []
    updated_punch_report = []
    index_report = 0
    for i in status_report:
        for j in punch_report:
            if j.get("emp_id") == i["emp_id"]:
                updated_punch_report.append(j)
                break
        for key, value in i.items():
            if type(value) == int:
                value = str(value)
            if value == None:
                value = ""
            if re.match(r'^[оуЯН ]+$', value):
                if "о" in value and "у" in value:
                    updated_punch_report[index_report][
                        key] = f'#E535FD {updated_punch_report[index_report][key]} #34A7FE'
                elif "о" in value:
                    updated_punch_report[index_report][
                        key] = f'#E535FD {updated_punch_report[index_report][key]}'
                elif "у" in value:
                    updated_punch_report[index_report][
                        key] = f'{updated_punch_report[index_report][key]} #34A7FE'
            if value == "Н":
                if updated_punch_report[index_report][key] != "":
                    updated_punch_report[index_report][key] = f'{updated_punch_report[index_report][key]} - Н'
                else:
                    updated_punch_report[index_report][key] = "Н"
            elif value == "О":
                updated_punch_report[index_report][key] = "О"
            elif value == "У":
                updated_punch_report[index_report][key] = "У"
            elif value == "Б":
                updated_punch_report[index_report][key] = "Б"
            elif value == "о Н" or value == "Н о":
                updated_punch_report[index_report][key] = f'{updated_punch_report[index_report][key]}-Н'
            elif value == 'K':
                updated_punch_report[index_report][key] = value
        index_report += 1
    punch_report = updated_punch_report
    punch_report = sorted(punch_report, key=lambda x: (x["first_name"] or ""))
    count_index = 0
    for i in punch_report:
        data_emp = []
        data_emp.append(i["first_name"])
        for index_punch, (key, value) in enumerate(i.items()):
            if index_punch > 29:
                if value == None:
                    value = '#002060 x'
                data_emp.append(value)
        data_emp.append(i["paycode_1"])
        data_emp.append(i["paycode_2"])
        data_emp.append(i["paycode_3"])
        if i["paycode_2"] != "":
            hours, minutes = i["paycode_2"].split(":")
            hours = int(hours)
            minutes = int(minutes)
            lateness_time = timedelta(hours=hours, minutes=minutes)
        else:
            lateness_time = ""
        if i["paycode_3"] != "":
            hours, minutes = i["paycode_3"].split(":")
            hours = int(hours)
            minutes = int(minutes)
            early_departure_time = timedelta(hours=hours, minutes=minutes)
        else:
            early_departure_time = ""
        if lateness_time != "" and early_departure_time != "":
            violations = lateness_time + early_departure_time
            violations = timedelta_hours_minutes_str(violations)
            data_emp.append(violations)
        elif lateness_time != "":
            violations = timedelta_hours_minutes_str(lateness_time)
            data_emp.append(violations)
        elif early_departure_time != "":
            violations = timedelta_hours_minutes_str(early_departure_time)
            data_emp.append(violations)
        else:
            data_emp.append("")
        data_emp.append(i["paycode_4"])
        data.append(data_emp)
        count_index = index_punch
    return data


def check_format(input_string):
    pattern = r'^\d{2}:\d{2} - Н$'
    return bool(re.match(pattern, input_string))


def create_header(ws, params):
    list_headers = [
        "Обычный", "Опоздание", "Уход раньше", "Нарушения", "Неявка",
    ]

    header = ["ИМЯ"]
    start_time = datetime.strptime(params["start_date"], "%Y-%m-%d")
    end_time = datetime.strptime(params["end_date"], "%Y-%m-%d")

    while True:
        header.append(f'{str(start_time.day)}.{str(start_time.month)}')
        if start_time.date() == end_time.date():  # - datetime.timedelta(days=1):
            break
        start_time += timedelta(days=1)
    header.extend(list_headers)
    ws.append(header)

    for cell in ws[1]:
        cell.font = Font(bold=True, color="ffffff")
        cell.alignment = Alignment(horizontal="center", vertical="center")
        cell.fill = PatternFill(start_color="4F81BD", end_color="4F81BD", fill_type="solid")
        cell.border = Border(
            left=Side(style='thin'),
            right=Side(style='thin'),
            top=Side(style='thin'),
            bottom=Side(style='thin')
        )


def fill_data(ws, data):
    for row in data:
        ws.append(row)

    fill_gray = PatternFill(start_color="DCE6F1", end_color="DCE6F1", fill_type="solid")
    fill_white = PatternFill(start_color="FFFFFF", end_color="FFFFFF", fill_type="solid")
    fill_red = PatternFill(start_color="FF0000", end_color="FF0000", fill_type="solid")
    fill_yellow = PatternFill(start_color="F9C01A", end_color="F9C01A", fill_type="solid")
    fill_purple = PatternFill(start_color="E535FD", end_color="E535FD", fill_type="solid")
    fill_blue = PatternFill(start_color="34A7FE", end_color="34A7FE", fill_type="solid")
    fill_orange = PatternFill(start_color="E26B0A", end_color="E26B0A", fill_type="solid")
    fill_siniy = PatternFill(start_color="002060", end_color="002060", fill_type="solid")

    for idx, row in enumerate(ws.iter_rows(min_row=2, max_col=len(data[0]), max_row=ws.max_row), start=2):
        fill = fill_gray if idx % 2 == 0 else fill_white
        for cell in row:
            if "#E535FD" in cell.value and "#34A7FE" in cell.value:
                f, time, s = cell.value.split(" ")
                cell.value = time
                cell.fill = fill_orange
            elif "#E535FD" in cell.value:
                f, time = cell.value.split(" ")
                cell.value = time
                cell.fill = fill_purple
            elif "#34A7FE" in cell.value:
                time, s = cell.value.split(" ")
                cell.value = time
                cell.fill = fill_blue
            elif "#002060" in cell.value:
                f, time = cell.value.split(" ")
                cell.value = time
                cell.fill = fill_siniy
            elif cell.value == "Н":
                cell.fill = fill_red
            elif check_format(cell.value):
                cell.fill = fill_yellow
            else:
                cell.fill = fill
            cell.border = Border(
                left=Side(style='thin'),
                right=Side(style='thin'),
                top=Side(style='thin'),
                bottom=Side(style='thin')
            )
            cell.alignment = Alignment(horizontal="center", vertical="center")
    for column in ws.columns:
        max_length = 0
        column_letter = get_column_letter(column[0].column)
        for cell in column:
            try:
                if cell.value and len(str(cell.value)) > max_length:
                    max_length = len(str(cell.value))
            except:
                pass
        adjusted_width = max_length + 2  # Добавляем небольшой запас
        ws.column_dimensions[column_letter].width = adjusted_width


class GetEmpReport(APIView):
    permission_classes = (IsAuthenticated,)
    accs = {
        "test": "3",
        "Anatoliy": "8",
        "au.cvetcov": "5",
    }

    def get(self, request):
        start_date = request.query_params.get("start_date")
        end_date = request.query_params.get("end_date")
        employees = request.query_params.get("employees")
        export = request.query_params.get("exp_data")
        if not start_date:
            return Response({"error": f"Нет параметра: start_date"}, status=status.HTTP_400_BAD_REQUEST)
        if not end_date:
            return Response({"error": f"Нет параметра: end_date"}, status=status.HTTP_400_BAD_REQUEST)
        end_date = datetime.strptime(end_date, "%Y-%m-%d")
        if end_date > datetime.now():
            end_date = datetime.now()
        if not request.user.is_staff:
            employees = self.accs[request.user.username]
        # Проверка разницы между end_time и start_time
        forty_days_ago = end_date - timedelta(days=40)
        start_date = datetime.strptime(start_date, "%Y-%m-%d")
        if start_date < forty_days_ago:
            start_date = forty_days_ago
        zk_api_interaction = ZkApiInteraction()
        zk_api_interaction.refresh_token()
        if isinstance(zk_api_interaction.token, dict) and 'status_code' in zk_api_interaction.token:
            content = zk_api_interaction.token
        else:
            if not employees:
                emp = zk_api_interaction.get_emp(user=request.user)
                employees = []
                for i in emp["data"]:
                    employees.append(i["id"])
                params = {"employees": ",".join(map(str, employees)), }
            else:
                params = {"employees": employees}
            start_date = datetime.strftime(start_date, "%Y-%m-%d")
            end_date = datetime.strftime(end_date, "%Y-%m-%d")
            params["start_date"] = start_date
            params["end_date"] = end_date
            params["page_size"] = 100
            params["exp_data"] = export
            punch_report = zk_api_interaction.get_monthly_punch_report(params=params, user=request.user)
            punch_report = punch_report["data"]
            status_report = zk_api_interaction.get_monthly_status_report(params=params, user=request.user)
            status_report = status_report["data"]

            data = pars_data(status_report=status_report, punch_report=punch_report)
            if export:
                wb = openpyxl.Workbook()
                ws = wb.active
                ws.title = "События за месяц"
                create_header(ws, params)
                fill_data(ws, data)
                ws.freeze_panes = 'B2'

                # Создаём буфер для сохранения файла
                file_stream = BytesIO()
                wb.save(file_stream)
                file_stream.seek(0)
                response = HttpResponse(file_stream,
                                        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
                response['Content-Disposition'] = 'attachment; filename="export.xlsx"'
                return response
            else:
                content = data
        return Response(content)


class Protected(APIView):
    permission_classes = (IsAuthenticated,)

    def get(self, request):
        return Response({"Test": "ok"}, status=status.HTTP_200_OK)


class ManualRecording(APIView):
    permission_classes = (IsAuthenticated,)

    def get(self, request):
        if request.user.is_superuser:
            return Response({"message": "Пользователь является суперпользователем."}, status=status.HTTP_200_OK)
        else:
            return Response({"message": "Недостаточно прав"},
                            status=status.HTTP_403_FORBIDDEN)
